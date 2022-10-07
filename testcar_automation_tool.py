import json
from tkinter import Y
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy
import os
import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic
import re
import webbrowser

path = os.getcwd()
file_path = path + "/config.json"
form_class = uic.loadUiType(path+"/view.ui")[0]

COLOR_INDEX = (
    '00000000', #black
    '00FFFFFF', #white
    '00FFFF00', #yellow
    '00FF0000', #red
    '00CCFFCC', #green
    '0016365C', #navi
    '00F79646'  #orange
)

COLOR_BLACK = COLOR_INDEX[0]
COLOR_WHITE = COLOR_INDEX[1]
COLOR_YELLOW = COLOR_INDEX[2]
COLOR_RED = COLOR_INDEX[3]
COLOR_GREEN = COLOR_INDEX[4]
COLOR_NAVI = COLOR_INDEX[5]
COLOR_ORANGE = COLOR_INDEX[6]

class MyWindow(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.connection()
        self.initialize()

        #data_dict format 
        # key   : (int)     index number
        # value : (string)  용도/차량정보/날짜,비용,수량,단위,소명여부 (ex. 주유비/1111(GN7)/'22.10.10,100000,1,EA,Y)
        self.data_dict = {}

    def show_popup(self, title, text):
        dlg = QMessageBox(self)
        dlg.setWindowTitle(title)
        dlg.setText(text)
        button = dlg.exec()

        if button == QMessageBox.Ok:
            print("OK!")

    def merge_purchase_cell(self, ws, start_idx):
        ws.merge_cells('B'+str(start_idx) + ":B"+str(start_idx+1))
        ws.merge_cells('C'+str(start_idx) + ":E"+str(start_idx))
        ws.merge_cells('F'+str(start_idx) + ":J"+str(start_idx))
        ws.merge_cells('K'+str(start_idx) + ":M"+str(start_idx))
        ws.merge_cells('D'+str(start_idx+1) + ":E"+str(start_idx+1))
        ws.merge_cells('F'+str(start_idx+1) + ":G"+str(start_idx+1))
        ws.merge_cells('H'+str(start_idx+1) + ":J"+str(start_idx+1))
        ws.merge_cells('K'+str(start_idx+1) + ":M"+str(start_idx+1))

    def merge_cell(self, ws):
        ws.merge_cells('A1:M1')

        ws.merge_cells('B5:C5')
        ws.merge_cells('D5:H5')
        ws.merge_cells('I5:J5')
        ws.merge_cells('K5:M5')

        ws.merge_cells('B6:C6')
        ws.merge_cells('D6:G6')
        ws.merge_cells('H6:I6')
        ws.merge_cells('J6:M6')

        ws.merge_cells('B7:C7')
        ws.merge_cells('D7:F7')
        ws.merge_cells('G7:H7')
        ws.merge_cells('I7:M7')

        ws.merge_cells('B8:C8')
        ws.merge_cells('D8:M8')

        ws.merge_cells('B9:M9')

        ws.merge_cells('B10:B11')
        ws.merge_cells('C10:E10')
        ws.merge_cells('D11:E11')

        ws.merge_cells('F10:J10')
        ws.merge_cells('F11:G11')
        ws.merge_cells('H11:J11')

        ws.merge_cells('K10:M10')
        ws.merge_cells('K11:M11')

    def set_request_info_cell(self, ws):
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col in range(1,30):
            ws.column_dimensions[get_column_letter(col)].width = 7
            ws.column_dimensions['A'].width = 2
            ws.column_dimensions['J'].width = 50
            ws.column_dimensions['N'].width = 2

        for row in range(5,100):
            ws.row_dimensions[row].height = 20

        date = data['date'] if data['date'] != "" else datetime.today().strftime("%Y.%m.%d")
        info_cell_dic = {'A1':'구매 의뢰 내역', 'I3':'청구번호:', 'J3':0, 'B5':'담당자', 
                        'D5':data['manager'], 'I5':'의뢰일', 'K5': date, 'B6':'프로젝트명', 'D6': data['project_name'], 'H6': '집행 부서',
                        'J6':data['team_name'], 'B7':'예산 NO', 'D7': data['budget_num'], 'G7':'예산명', 'I7': data['budget_name'], 
                        'B8': '용도', 'D8': data['purpose']}

        for key, val in info_cell_dic.items():
            ws[key] = val
            info_cell = ws[key]

            if key == "A1":
                info_cell.font = Font(color=COLOR_BLACK, bold=True, underline='single', size=18, name='굴림')
                info_cell.alignment = Alignment(horizontal='center')
            elif key == "I3" or key == "J3":
                info_cell.font = Font(size=11, bold=True, name='현대하모니 L')
                info_cell.fill = PatternFill(fgColor=COLOR_YELLOW, fill_type='solid')
                if key == "J3":
                    info_cell.alignment = Alignment(horizontal='right')
            elif key == "D5" or key == "D6" or key == "I7":
                info_cell.font = Font(size=10, name='현대하모니 L')
                info_cell.alignment = Alignment(horizontal='center', vertical='center')
                info_cell.border = thin_border
            else:
                info_cell.font = Font(size=12, name='현대하모니 L')
                info_cell.alignment = Alignment(horizontal='center', vertical='center')
                if key == "I5" or key == "H6" or key == "B6" or key == "B7" or key == "G7" or key == "B8" or key == "B5":
                    info_cell.fill = PatternFill(fgColor=COLOR_YELLOW, fill_type='solid')
                info_cell.border = thin_border

        common_cell_dic = {'B9':'구 매      내 역', 'B10':'NO', 'C10':'품 번', 'C11':'단위', 'D11':'단 가', 'F10':'품 명', 'K10':'추천 업체',
                            'F11':'신청량', 'H11': '구입가', 'K11': '입고 요구일'}

        for key, val in common_cell_dic.items():
            ws[key] = val
            common_cell = ws[key]
            common_cell.font = Font(size=12, name='현대하모니 L')
            common_cell.fill = PatternFill(fgColor=COLOR_YELLOW, fill_type='solid')
            common_cell.alignment = Alignment(horizontal='center', vertical='center')
            common_cell.border = thin_border
    
        #병합 한 이후 style을 주면 border 처리에 오류가 발생함. 
        #따라서 style을 모두 주고 마지막에 병합처리할 경우 border가 정상적으로 처리됨... opensource의 문제점...ㅠ.ㅠ
        self.merge_cell(ws)
    
    def set_purchase_info_cell(self, ws):
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        start_row = 12

        for key, val in self.data_dict.items():
            index_cell_id = 'B' + str(start_row)
            ws[index_cell_id] = key
            index_cell = ws[index_cell_id]
            index_cell.font = Font(size=12, name='현대하모니 L')
            index_cell.alignment = Alignment(horizontal='center', vertical='center')
            index_cell.border = thin_border
            index_cell.fill = PatternFill(fgColor=COLOR_YELLOW, fill_type='solid')

            split_data = val.split(",")
            total_cost = int(split_data[1]) * int(split_data[2])
            cost_cell_id = 'D' + str(start_row+1)
            ws[cost_cell_id] = "=TEXT("+str(total_cost)+",\"₩#,##0\")"
            cost_cell = ws[cost_cell_id]
            cost_cell.font = Font(size=12, name='현대하모니 L')
            cost_cell.alignment = Alignment(horizontal='center', vertical='center')
            cost_cell.border = thin_border

            discription_cell_id = 'F' + str(start_row)
            ws[discription_cell_id] = split_data[0]
            discription_cell = ws[discription_cell_id]
            if(split_data[4] == 'Y'):
                discription_cell.font = Font(size=11, bold=True, name='현대하모니 L')
                discription_cell.fill = PatternFill(fgColor=COLOR_ORANGE, fill_type='solid')
            else:
                discription_cell.font = Font(size=11, name='현대하모니 L')
            
            discription_cell.alignment = Alignment(horizontal='center', vertical='center')
            discription_cell.border = thin_border

            #빈칸 색 채우기
            common_cell = ws['C'+str(start_row+1)]
            common_cell.border = thin_border
            common_cell.fill = PatternFill(fgColor=COLOR_GREEN, fill_type='solid')

            common_cell = ws['H'+str(start_row+1)]
            common_cell.border = thin_border
            common_cell.fill = PatternFill(fgColor=COLOR_GREEN, fill_type='solid')

            common_cell = ws['K'+str(start_row+1)]
            common_cell.border = thin_border
            common_cell.fill = PatternFill(fgColor=COLOR_NAVI, fill_type='solid')

            common_cell = ws['C'+str(start_row)]
            common_cell.border = thin_border
            common_cell = ws['K'+str(start_row)]
            common_cell.border = thin_border
            common_cell = ws['F'+str(start_row+1)]
            common_cell.border = thin_border

            self.merge_purchase_cell(ws, start_row)
            start_row += 2

    def export_excel_files(self):
        self.export_purchase_order_excel()
        self.export_helper_excel()

    def export_purchase_order_excel(self):
        wb = Workbook()
        work_sheet = wb.active
        work_sheet.title = "시험차"

        #기본 정보 설정
        self.set_request_info_cell(work_sheet)
        self.set_purchase_info_cell(work_sheet)
        file_title = str(datetime.today().year)+"년_"+str(datetime.today().month)+"월_"+"구매의뢰.xlsx"
        try:
            wb.save(path + '/' + file_title)
            wb.close()
            self.show_popup("Completed!!", "Excel Export가 완료되었습니다.")
        except PermissionError:
            self.show_popup("Warning", file_title +" Excel file이 열려있습니다.")

    def export_helper_excel(self):
        wb = Workbook()
        work_sheet = wb.active
        total_cost = 0
        
        for key, val in self.data_dict.items():
            cost = 1
            #index 추가
            work_sheet['A' + str(key)] = key

            split_data = val.split(",")
            for index in range(1, len(split_data)):
                cell_idx = get_column_letter(index+1) + str(key)

                if index == 1:
                    work_sheet.column_dimensions[get_column_letter(index+1)].width = 30

                if split_data[index-1].isdigit():
                    work_sheet[cell_idx] = int(split_data[index-1])
                    cost *= int(split_data[index-1])
                else:
                    work_sheet[cell_idx] = split_data[index-1]
    
            total_cost += cost
        
        work_sheet['F1'] = total_cost
        cell = work_sheet['F1']
        cell.font = Font(size=12, name='현대하모니 L', color=COLOR_RED)
        cell.fill = PatternFill(fgColor=COLOR_YELLOW, fill_type='solid')

        try:
            wb.save(path + '/'+str(datetime.today().month)+'월_'+'시스템추가필요.xlsx')
            wb.close()
        except PermissionError:
            print(str(datetime.today().month)+'월_'+'시스템추가필요 Excel file이 열려있습니다.')

    def get_basic_info_config(self):
        with open(file_path, 'r', encoding="UTF-8") as fp:
            global data
            _info = json.load(fp)
            data = _info["basic_info"]

    def get_car_info_config(self):
        with open(file_path, 'r', encoding="UTF-8") as fp:
            global car_info
            _info = json.load(fp)
            car_info = _info["car_info"]

    def set_car_info_combobox(self):
        for key, value in car_info.items() :
            self.car_num_combo.addItem(key)
    
    def check_date_format(self, txt):
        try:
            regex = r'\d{2}.\d{2}.\d{2}'
            return  bool(re.match(regex, txt))
        except ValueError:
            self.show_popup("Warning", "잘못된 날짜 포멧입니다.\n YY.MM.DD포멧으로 입력하세요.")
            return False

    def add_purchasing_info(self):
        if self.cost_edit.text() == "":
            self.show_popup("Warning", "비용을 추가하세요.")
        elif self.description_edit.text() == "":
            self.show_popup("Warning", "사용처를 추가하세요.")
        elif self.date_edit.text() == "":
            self.show_popup("Warning", "날짜를 추가하세요.")
        elif self.count_edit.text() == "":
            self.show_popup("Warning", "수량을 추가하세요.")
        else:
            if not self.cost_edit.text().isdigit():
                self.show_popup("Warning", "잘못된 비용 포멧입니다.\n 비용은 숫자로만 입력하세요.")
                return
            if self.check_date_format(self.date_edit.text()):
                car_num = self.car_num_combo.currentText()
                index_num = self.list_widget.count() + 1
                add_str = str(index_num) + ".   " + self.description_edit.text() + " / " + car_num + " (" + car_info[car_num] + ") / '" + self.date_edit.text()
                self.data_dict[index_num] = self.description_edit.text() + " / " + car_num + " (" + car_info[car_num] + ") / '" + self.date_edit.text() + "," + self.cost_edit.text()+","+self.count_edit.text()+",EA,"+self.check_combo.currentText()
                self.list_widget.addItem(add_str)
            else:
                self.show_popup("Warning", "잘못된 날짜 포멧입니다.\n YY.MM.DD포멧으로 입력하세요.")

    def remove_dic_data(self, key):
        r = dict(self.data_dict)
        del r[key]
        return r

    def remove_current_item(self):
        removeItemRow = self.list_widget.currentRow()
        self.remove_dic_data(removeItemRow+1)
        self.list_widget.takeItem(removeItemRow)

    def clear_all_items(self):
        self.list_widget.clear()
        self.data_dict = {}

    def launch_user_guide(self):
        url = "https://synergy.ccos.dev/pages/viewpage.action?pageId=115448499"
        webbrowser.open(url)

    def show_about_helper(self):
        about_helper_str = "시험차 전표처리 Helper tool\n - version : 0.1\n - compile date : "+ str(datetime.today().strftime("%Y.%m.%d")) + "\n\n provide by kyungjoon.hyun"
        self.show_popup("About Helper", about_helper_str)

    def connection(self):
        self.add_button.clicked.connect(self.add_purchasing_info)
        self.remove_button.clicked.connect(self.remove_current_item)
        self.clear_button.clicked.connect(self.clear_all_items)
        self.export_button.clicked.connect(self.export_excel_files)
        self.actionUser_Guide.triggered.connect(self.launch_user_guide)
        self.actionAbout_helper.triggered.connect(self.show_about_helper)

    def initialize(self):
        self.get_basic_info_config()
        self.get_car_info_config()
        self.set_car_info_combobox()


if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()
